from flask import (Flask, redirect, render_template, request,
                   send_from_directory, url_for, abort, Response)

import pandas as pd
import numpy as np
from dotenv import load_dotenv
import os

import hashlib
import hmac
from termcolor import colored


import msal
import requests
import time
from datetime import datetime

import threading

import os
import shutil
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader

from time import sleep

load_dotenv()

CLIENT_ID = os.getenv("CLIENT_ID")
CLIENT_SECRET = os.getenv("CLIENT_SECRET")
TENANT_ID = os.getenv("TENANT_ID")
DRIVE_ID = os.getenv("DRIVE_ID")    

FILE_PATH = "Valinta-apuri/data.xlsx"
OUTPUT_FILE = "dp/d.xlsx"
POLL_INTERVAL_SECONDS = 300

AUTHORITY = f"https://login.microsoftonline.com/{TENANT_ID}"

# --------- MSAL CLIENT ---------
msal_app = msal.ConfidentialClientApplication(
    client_id=CLIENT_ID,
    client_credential=CLIENT_SECRET,
    authority=AUTHORITY
)

def get_token():
    return msal_app.acquire_token_for_client(
        scopes=["https://graph.microsoft.com/.default"]
    ).get("access_token")

def download_file():
    headers = {"Authorization": f"Bearer {get_token()}"}
    r = requests.get(
        f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/root:/{FILE_PATH}",
        headers=headers
    )
    if r.status_code != 200:
        print(f"❌ Get file failed: {r.status_code}")
        return
    download_url = r.json()["@microsoft.graph.downloadUrl"]
    r = requests.get(download_url)
    if r.status_code == 200:
        with open(OUTPUT_FILE, "wb") as f:
            f.write(r.content)
        print(f"✅ Downloaded '{OUTPUT_FILE}' ({len(r.content):,} bytes)")

def get_last_modified():
    headers = {"Authorization": f"Bearer {get_token()}"}
    r = requests.get(
        f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/root:/{FILE_PATH}",
        headers=headers
    )
    if r.status_code != 200:
        print(f"❌ Check failed: {r.status_code}")
        return None
    return r.json().get("lastModifiedDateTime")

def get_file_hashes():
    headers = {"Authorization": f"Bearer {get_token()}"}
    r = requests.get(
        f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/root:/{FILE_PATH}",
        headers=headers
    )
    if r.status_code != 200:
        print(f"❌ Check failed: {r.status_code}")
        return None
    
    hashes = r.json().get("file", {}).get("hashes", {})
    print(f"   quickXorHash : {hashes.get('quickXorHash')}")
    print(f"   sha256Hash   : {hashes.get('sha256Hash')}")
    return hashes.get("quickXorHash") or hashes.get("sha256Hash")


def poller():
    print("🚀 Starting poller (every 5 minutes)")

    download_file()
    last_hash = get_file_hashes()
    print(f"📄 Initial hash: {last_hash}")

    while True:
        time.sleep(POLL_INTERVAL_SECONDS)
        global df, categories 
        print(f"\n🔍 Checking hash... ({datetime.now().strftime('%H:%M:%S')})")

        current_hash = get_file_hashes()

        if current_hash and current_hash != last_hash:
            print(f"🔔 Hash changed! Downloading...")
            download_file()
            last_hash = current_hash
            df, categories = load_and_process_excel()
            load_img_from_excel()
        else:
            print(f"   No changes.")


def strip_list(items):
    """Trim whitespace for each item and normalize capitalization.

    This makes the first character uppercase and the rest lowercase (e.g. "ma" -> "Ma").
    It is used for columns like `Days` and `Category` so values match the UI labels.
    """
    out = []
    for item in items:
        s = str(item).strip()
        if s:
            s = s.capitalize()
        out.append(s)
    return out

def to_int_list(items):
    return [int(item.strip()) for item in items if item.strip().isdigit()]


def load_and_process_excel(file_path='dp/d.xlsx'):
    """
    Load and process Excel file data.
    
    Args:
        file_path (str): Path to the Excel file
        
    Returns:
        tuple: (processed_dataframe, categories_list)
    """
    # Read the Excel file, using the first row as column headers
    df = pd.read_excel(file_path, header=0)

    for col in ['Days', 'Category']:
        if col in df.columns:
            df[col] = df[col].astype(str).str.split(',')
            df[col] = df[col].apply(strip_list)

    if 'Level' in df.columns:
        df['Level'] = df['Level'].astype(str).str.split(',')
        df['Level'] = df['Level'].apply(to_int_list)

    df['Category'] = df['Category'].replace('', np.nan)
    df['Location'] = df['Location'].replace('', np.nan)
    cdf = df.dropna(subset=['Category'])
    categories = (
        cdf['Category']
        .explode()
        .map(lambda x: x.strip())
        .dropna()
        .loc[lambda s: (s != '') & (s.str.lower() != 'nan')]
        .unique()
        .tolist()
    )
    
    return df, categories

        
        
def load_img_from_excel():
    tmp_folder = "static/img_tmp"
    cur_folder = "static/img_cur"

    # Work in tmp folder (clean slate)
    if os.path.exists(tmp_folder):
        shutil.rmtree(tmp_folder)
    os.makedirs(tmp_folder, exist_ok=True)

    # Ensure cur folder exists
    os.makedirs(cur_folder, exist_ok=True)

    wb = load_workbook('dp/d.xlsx')
    sheet = wb['Sheet1']
    image_loader = SheetImageLoader(sheet)

    # Process and save all images to tmp folder first
    for row in range(2, sheet.max_row + 1):
        cell = f'H{row}'
        img_name = row - 2
        if image_loader.image_in(cell):
            image = image_loader.get(cell)
            image.save(os.path.join(tmp_folder, f'{img_name}.png'))
            print(f"Saved image in {cell} as {tmp_folder}/{img_name}.png")

    # Swap: clear cur and copy tmp -> cur (minimizes downtime)
    if os.path.exists(cur_folder):
        shutil.rmtree(cur_folder)
    shutil.copytree(tmp_folder, cur_folder)

    print("✅ Pictures reloaded successfully!")
    

# Start the polling in a separate thread so it doesn't block the Flask app
poller_thread = threading.Thread(target=poller, daemon=True)
poller_thread.start()

time.sleep(5) # Wait a bit before loading images to ensure the first download is complete


# Load initial data
df, categories = load_and_process_excel()
load_img_from_excel()

app = Flask(__name__)


@app.route('/', methods=['GET'])
def index():
    # Collect selected levels, days, and categories from query parameters

    selected_days = [day for day in ["Ma", "Ti", "Ke", "To", "Pe"] if request.args.get(day) == "True"]
    
    # Collect individual lvlN keys (if present) and group values submitted under 'lvl_group'
    selected_levels = set(
        int(key.replace('lvl', ''))
        for key in request.args
        if key.startswith('lvl') and request.args.get(key) == "True"
    )

    # request.args.getlist('lvl_group') returns multiple selected group values (e.g. "1,2")
    for group_val in request.args.getlist('lvl_group'):
        for part in group_val.split(','):
            part = part.strip()
            if part.isdigit():
                selected_levels.add(int(part))

    # Use a sorted list downstream (same shape as before)
    selected_levels = sorted(selected_levels)
    
    selected_locations = [loc for loc in df['Location'].unique() if request.args.get(loc) == "True"]

    # Debugging help: print raw and parsed values so we can spot encoding/parse issues
    try:
        print("DEBUG request.args:", dict(request.args))
        print("DEBUG lvl_group raw:", request.args.getlist('lvl_group'))
        print("DEBUG selected_levels (after parse):", selected_levels)
    except Exception:
        pass
    
    selected_categories = [cat for cat in categories if request.args.get(cat) == "True"]

    # Show no data if no filters are selected
    if not selected_levels and not selected_days and not selected_categories:
        df_filtered = df.copy() #df.iloc[0:0]  # Empty DataFrame with same columns
    else:
        df_filtered = df.copy()
        # Apply filters in the order: Category -> Level -> Day
        if selected_categories and not df_filtered.empty:
            df_filtered = df_filtered[df_filtered['Category'].apply(lambda cats: any(cat in cats for cat in selected_categories))]
        if selected_levels and not df_filtered.empty:
            df_filtered = df_filtered[df_filtered['Level'].apply(lambda levels: any(lvl in levels for lvl in selected_levels))]
        if selected_days and not df_filtered.empty:
            df_filtered = df_filtered[df_filtered['Days'].apply(lambda days: any(day in days for day in selected_days))]
        if selected_locations and not df_filtered.empty:
            df_filtered = df_filtered[df_filtered['Location'].apply(lambda locs: any(loc in str(locs) for loc in selected_locations))]
            
    return render_template(
      'index.html',
      # Pass a concrete list so the value can be iterated multiple times in templates
      aste=list(enumerate(["1. Luokka", "2. Luokka", "3. Luokka", "4. Luokka", "5. Luokka",
          "6. Luokka", "7. Luokka", "8. Luokka", "9. Luokka", "2. Aste"])),
        selected_levels=selected_levels,
        selected_days=selected_days,
        selected_categories=selected_categories,
        selected_locations=selected_locations,
        args=request.args,
        df=df_filtered.to_html(classes='data', header="true", index=True, justify='center'),
        rowItems=df_filtered.itertuples(name=None),
        categories=categories,
        locations=df['Location'].unique()
    )

@app.route("/test", methods=['GET'])
def test():
    return df.to_html(classes='data', header="true")
    # return render_template('test.html')
    
    

    
@app.errorhandler(404)
def page_not_found(e):
    # Check if the request was for a missing image in img_tmp
    # print(request.headers.get('Accept'))
    if request.path.startswith('/static/img_tmp/') and request.path.endswith('.png'):
        # Return a generic image from your static folder
        return send_from_directory('static', 'generic.jpg'), 307
    # Otherwise, show the normal 404 page
    return render_template('404.html'), 404

if __name__ == '__main__':
    app.run(debug=False, host="0.0.0.0", port=80)