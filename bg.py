from flask import (Flask, redirect, render_template, request,
                   send_from_directory, url_for, abort, Response)

import pandas as pd
import numpy as np
from dotenv import load_dotenv
import os

import hashlib
import hmac
from termcolor import colored


import msal
import requests
import time
from datetime import datetime

import threading

import os
import shutil
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader

from time import sleep

from PIL import Image
import filetype



# If not in production, load .env for development. In production, we expect environment variables from systemd.
if os.getenv("ENV") != "production":
    from dotenv import load_dotenv
    load_dotenv()


CLIENT_ID = os.getenv("CLIENT_ID")
CLIENT_SECRET = os.getenv("CLIENT_SECRET")
TENANT_ID = os.getenv("TENANT_ID")
DRIVE_ID = os.getenv("DRIVE_ID")    

FILE_PATH = "Valinta-apuri/data.xlsx"
OUTPUT_FILE = "dp/d.xlsx"
POLL_INTERVAL_SECONDS = 180 # 3 minutes

AUTHORITY = f"https://login.microsoftonline.com/{TENANT_ID}"

# --------- MSAL CLIENT ---------
msal_app = msal.ConfidentialClientApplication(
    client_id=CLIENT_ID,
    client_credential=CLIENT_SECRET,
    authority=AUTHORITY
)

def get_token():
    return msal_app.acquire_token_for_client(
        scopes=["https://graph.microsoft.com/.default"]
    ).get("access_token")

def download_file():
    headers = {"Authorization": f"Bearer {get_token()}"}
    r = requests.get(
        f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/root:/{FILE_PATH}",
        headers=headers
    )
    if r.status_code != 200:
        print(f"❌ Get file failed: {r.status_code}")
        return
    download_url = r.json()["@microsoft.graph.downloadUrl"]
    r = requests.get(download_url)
    if r.status_code == 200:
        with open(OUTPUT_FILE, "wb") as f:
            f.write(r.content)
        print(f"✅ Downloaded '{OUTPUT_FILE}' ({len(r.content):,} bytes)")

def get_last_modified():
    headers = {"Authorization": f"Bearer {get_token()}"}
    r = requests.get(
        f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/root:/{FILE_PATH}",
        headers=headers
    )
    if r.status_code != 200:
        print(f"❌ Check failed: {r.status_code}")
        return None
    return r.json().get("lastModifiedDateTime")

def get_file_hashes():
    headers = {"Authorization": f"Bearer {get_token()}"}
    r = requests.get(
        f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/root:/{FILE_PATH}",
        headers=headers
    )
    if r.status_code != 200:
        print(f"❌ Check failed: {r.status_code}")
        return None
    
    hashes = r.json().get("file", {}).get("hashes", {})
    print(f"   quickXorHash : {hashes.get('quickXorHash')}")
    print(f"   sha256Hash   : {hashes.get('sha256Hash')}")
    return hashes.get("quickXorHash") or hashes.get("sha256Hash")

def strip_list(items):
    """Trim whitespace for each item and normalize capitalization.

    This makes the first character uppercase and the rest lowercase (e.g. "ma" -> "Ma").
    It is used for columns like `Days` and `Category` so values match the UI labels.
    """
    out = []
    for item in items:
        s = str(item).strip()
        if s:
            s = s.capitalize()
        out.append(s)
    return out

def to_int_list(items):
    return [int(item.strip()) for item in items if item.strip().isdigit()]


def load_and_process_excel(file_path='dp/d.xlsx'):
    """
    Load and process Excel file data.
    
    Args:
        file_path (str): Path to the Excel file
        
    Returns:
        tuple: (processed_dataframe, categories_list)
    """
    # Read the Excel file, using the first row as column headers
    df = pd.read_excel(file_path, header=0)

    for col in ['Days', 'Category']:
        if col in df.columns:
            df[col] = df[col].astype(str).str.split(',')
            df[col] = df[col].apply(strip_list)

    if 'Level' in df.columns:
        df['Level'] = df['Level'].astype(str).str.split(',')
        df['Level'] = df['Level'].apply(to_int_list)

    df['Category'] = df['Category'].replace('', np.nan)
    df['Location'] = df['Location'].replace('', np.nan)
    df = df.dropna(subset=['Workshop', 'Location'])
    cdf = df.dropna(subset=['Category'])
    categories = (
        cdf['Category']
        .explode()
        .map(lambda x: x.strip())
        .dropna()
        .loc[lambda s: (s != '') & (s.str.lower() != 'nan')]
        .unique()
        .tolist()
    )
    
    
    return df, pd.Series(categories)

def load_img_from_excel():
    tmp_folder = "static/img_tmp"
    cur_folder = "static/img_cur"

    # Work in tmp folder (clean slate)
    if os.path.exists(tmp_folder):
        shutil.rmtree(tmp_folder)
    os.makedirs(tmp_folder, exist_ok=True)

    # Ensure cur folder exists
    os.makedirs(cur_folder, exist_ok=True)

    wb = load_workbook('dp/d.xlsx')
    sheet = wb['Sheet1']
    image_loader = SheetImageLoader(sheet)

    # Process and save all images to tmp folder first
    for row in range(2, sheet.max_row + 1):
        cell = f'H{row}'
        img_name = row - 2
        if image_loader.image_in(cell):
            image = image_loader.get(cell)
            image.save(os.path.join(tmp_folder, f'{img_name}.png'))
            print(f"Saved image in {cell} as {tmp_folder}/{img_name}.png")
            
    # Convert PNGs to WEBPs and resize (to optimize for web)
    for img_file in os.listdir(tmp_folder):
        img_path = os.path.join(tmp_folder, img_file)
        if os.path.isfile(img_path) and filetype.is_image(img_path):
            with Image.open(img_path) as img:
                webp_path = img_path.rsplit('.', 1)[0] + '.webp'
                w, h = img.size
                new_h = 400
                new_w = int(w * (new_h / h))
                img = img.resize((new_w, new_h))
                img.convert('RGB').save(webp_path, format='WEBP', quality=80, method=6)
                os.remove(img_path)  # Remove original PNG
                print(f"Converted {img_file} to {webp_path}")
    

    # Swap: clear cur and copy tmp -> cur (minimizes downtime)
    if os.path.exists(cur_folder):
        shutil.rmtree(cur_folder)
    shutil.copytree(tmp_folder, cur_folder)

    print("✅ Pictures reloaded successfully!")
    

def poller():
    print("🚀 Starting poller")

    
    last_hash = "asdasdasd" # Dummy initial value to ensure the first check triggers a download
    

    while True:
        global df, categories 
        print(f"\n🔍 Checking hash... ({datetime.now().strftime('%H:%M:%S')})")

        current_hash = get_file_hashes()

        if current_hash and current_hash != last_hash:
            print(f"🔔 Hash changed! Downloading...")
            try:
                download_file()
                last_hash = current_hash
                print(f"Download successful, updated hash: {last_hash}")
            except Exception as e:
                print(f"❌ Download failed: {e}")
                continue
            
            try:
                df, categories = load_and_process_excel()
                print(f"✅ Excel data reloaded successfully!")
                df.to_pickle("./tmp/data.pkl")
                categories.to_pickle("./tmp/categories.pkl")
            except Exception as e:
                print(f"❌ Excel processing failed: {e}")
                continue
            
            try:
                load_img_from_excel()
                print(f"✅ Images reloaded successfully!")
            except Exception as e:
                print(f"❌ Image loading failed: {e}")
                continue
        else:
            print(f"   No changes.")
            
        time.sleep(POLL_INTERVAL_SECONDS)
            
            
if __name__ == "__main__":
    poller()